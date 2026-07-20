# -*- coding: utf-8 -*-
"""阶段0：读入题目清单(xlsx/csv) → manifest.json。
用法: python 00_import_list.py <输入.xlsx|.csv> [题目列序号(1起,默认自动探测)] [表头行(默认自动)]"""
import sys, os, json, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from lib import CFG, save_manifest, find_header
sys.stdout.reconfigure(encoding='utf-8')

def read_xlsx(path):
    import openpyxl
    ws = openpyxl.load_workbook(path, data_only=True).active
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]

def read_csv(path):
    import csv
    with open(path, encoding='utf-8-sig', newline='') as f:
        return [row for row in csv.reader(f)]

def main():
    if len(sys.argv) < 2:
        print("用法: python 00_import_list.py <输入.xlsx|.csv> [题目列(1起)] [表头行(1起)]"); return
    path = sys.argv[1]
    rows = read_csv(path) if path.lower().endswith(('.csv', '.tsv')) else read_xlsx(path)
    if len(sys.argv) >= 4:
        hdr = int(sys.argv[3]) - 1; tcol = int(sys.argv[2]) - 1; ncol = None
    else:
        hdr, tcol, ncol = find_header(rows)
        if len(sys.argv) >= 3: tcol = int(sys.argv[2]) - 1
    # 猜其他列
    header = rows[hdr]
    def col_of(*keys):
        for ci, v in enumerate(header):
            if v and any(k in str(v) for k in keys): return ci
        return None
    if ncol is None: ncol = col_of('序号', '编号', 'No', 'ID')
    ycol = col_of('年', 'year', 'Year'); tycol = col_of('类型', 'type')
    m = []
    n = 0
    for row in rows[hdr + 1:]:
        title = row[tcol] if tcol < len(row) else None
        if not title or not str(title).strip(): continue
        n += 1
        num = row[ncol] if (ncol is not None and ncol < len(row) and row[ncol]) else n
        try: num = int(num)
        except: num = n
        rec = {'num': num, 'excel_title': str(title).strip()}
        if ycol is not None and ycol < len(row) and row[ycol]: rec['excel_year'] = row[ycol]
        if tycol is not None and tycol < len(row) and row[tycol]: rec['excel_type'] = row[tycol]
        m.append(rec)
    save_manifest(m)
    print(f"导入 {len(m)} 篇 -> {CFG['manifest']}  (表头行={hdr+1}, 题目列={tcol+1})")

if __name__ == '__main__':
    main()
